"""planner/utils/export_import_utils.py — Semester export / import (JSON, XLSX, CSV)."""

import csv
import json
import re

# ── Column layout ─────────────────────────────────────────────────────────────
XLSX_HEADERS = [
    "Course Name", "Base Module", "Specific Module", "Credits",
    "Exam Given", "Exam Date", "Exam Time",
    "Alt Date",   "Alt Time",   "Additional Info", "Color",
]

COURSE_FIELD_MAP = [
    ("name",            "Course Name"),
    ("base_module",     "Base Module"),
    ("specific_module", "Specific Module"),
    ("credits",         "Credits"),
    ("exam_given",      "Exam Given"),
    ("exam_date",       "Exam Date"),
    ("exam_time",       "Exam Time"),
    ("alt_date",        "Alt Date"),
    ("alt_time",        "Alt Time"),
    ("additional_info", "Additional Info"),
    ("color",           "Color"),
]

# Catppuccin Mocha palette (for xlsx styling)
_BG       = "1E1E2E"
_CRUST    = "11111B"
_SURFACE0 = "313244"
_SURFACE1 = "45475A"
_FG       = "CDD6F4"
_SUBTEXT  = "A6ADC8"
_ACCENT   = "89B4FA"
_GREEN    = "A6E3A1"
_YELLOW   = "F9E2AF"
_RED      = "F38BA8"
_OVERLAY  = "7F849C"

_TEMPLATE_COURSES = [
    {
        "name": "Example Course A",
        "base_module": "Mathematics",
        "specific_module": "Analysis",
        "credits": 5,
        "exam_given": False,
        "exam_date": "2026-07-15",
        "exam_time": "10:00",
        "alt_date": "2026-09-10",
        "alt_time": "10:00",
        "additional_info": "Written exam — 90 min",
        "color": "#89B4FA",
    },
    {
        "name": "Example Course B",
        "base_module": "Computer Science",
        "specific_module": "Algorithms",
        "credits": 8,
        "exam_given": True,
        "exam_date": "2026-07-20",
        "exam_time": "14:00",
        "alt_date": "",
        "alt_time": "",
        "additional_info": "",
        "color": "#A6E3A1",
    },
    {
        "name": "Example Course C — delete & replace",
        "base_module": "Physics",
        "specific_module": "",
        "credits": 3,
        "exam_given": False,
        "exam_date": "",
        "exam_time": "",
        "alt_date": "",
        "alt_time": "",
        "additional_info": "No exam — graded coursework",
        "color": "#FAB387",
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# JSON
# ─────────────────────────────────────────────────────────────────────────────

def export_semester_json(semester: dict, path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(semester, f, indent=2, ensure_ascii=False)


def import_semester_json(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    # Accept either a bare semester dict or a full data.json (pick first semester)
    if "semesters" in data:
        sems = data["semesters"]
        if not sems:
            raise ValueError("JSON file contains no semesters.")
        return sems[0]
    if "name" in data and "courses" in data:
        return data
    raise ValueError("Unrecognised JSON structure — expected a semester object.")


# ─────────────────────────────────────────────────────────────────────────────
# CSV
# ─────────────────────────────────────────────────────────────────────────────

def export_semester_csv(semester: dict, path: str):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["# Semester",
                     semester.get("name", ""),
                     semester.get("display_name", "")])
        w.writerow([])
        w.writerow(XLSX_HEADERS)
        for course in semester.get("courses", []):
            row = []
            for field, _ in COURSE_FIELD_MAP:
                v = course.get(field, "")
                if isinstance(v, bool):
                    v = "TRUE" if v else "FALSE"
                row.append("" if v is None else v)
            w.writerow(row)


def import_semester_csv(path: str) -> dict:
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    meta = {"name": "", "display_name": ""}
    header_idx = None

    for i, row in enumerate(rows):
        if row and row[0] == "# Semester":
            meta["name"]         = row[1] if len(row) > 1 else ""
            meta["display_name"] = row[2] if len(row) > 2 else ""
        if row and row[0] == XLSX_HEADERS[0]:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError("Could not find header row — expected first column: 'Course Name'.")

    col_map = {h: i for i, h in enumerate(rows[header_idx])}
    courses = []

    for row in rows[header_idx + 1:]:
        if not row or not any(c.strip() for c in row):
            continue
        course = _parse_course_row(row, col_map)
        courses.append(course)

    return {
        "name":         meta["name"],
        "display_name": meta["display_name"],
        "courses":      courses,
        "exams":        [],
    }


def _parse_course_row(row, col_map) -> dict:
    course = {}
    for field, col_name in COURSE_FIELD_MAP:
        idx = col_map.get(col_name)
        val = (row[idx] if idx is not None and idx < len(row) else "") or ""
        if field == "exam_given":
            val = str(val).strip().upper() in ("TRUE", "1", "YES")
        elif field == "credits":
            try:
                val = int(float(str(val)))
            except (ValueError, TypeError):
                val = 0
        else:
            val = "" if str(val).strip() in ("None", "") else str(val)
        course[field] = val
    course.setdefault("slots", [])
    return course


# ─────────────────────────────────────────────────────────────────────────────
# XLSX
# ─────────────────────────────────────────────────────────────────────────────

def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX support.\n"
            "Install it with:  pip install openpyxl"
        )


def export_semester_xlsx(semester: dict, path: str):
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses"

    thin_border = Border(
        left=Side(style="thin",   color="45475A"),
        right=Side(style="thin",  color="45475A"),
        top=Side(style="thin",    color="45475A"),
        bottom=Side(style="thin", color="45475A"),
    )
    thick_bottom = Border(
        left=Side(style="thin",   color="45475A"),
        right=Side(style="thin",  color="45475A"),
        top=Side(style="thin",    color="45475A"),
        bottom=Side(style="medium", color=_ACCENT),
    )

    # ── Row 1: Semester title ─────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    sem_disp = semester.get("display_name") or semester.get("name", "")
    sem_short = semester.get("name", "")
    c.value = f"  {sem_disp}    ·    short name: {sem_short}"
    c.font      = Font(name="Arial", size=13, bold=True, color=_ACCENT)
    c.fill      = PatternFill("solid", fgColor=_CRUST)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # ── Row 2: blank spacer ───────────────────────────────────────────────────
    ws.append([])
    ws.row_dimensions[2].height = 4
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor=_CRUST)

    # ── Row 3: column headers ─────────────────────────────────────────────────
    ws.append(XLSX_HEADERS)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font      = Font(name="Arial", size=10, bold=True, color=_FG)
        cell.fill      = PatternFill("solid", fgColor=_SURFACE0)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = thick_bottom
    ws.row_dimensions[hdr_row].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, course in enumerate(semester.get("courses", [])):
        row_data = []
        for field, _ in COURSE_FIELD_MAP:
            v = course.get(field, "")
            if field == "exam_given":
                v = "TRUE" if v else "FALSE"
            elif field == "credits":
                v = int(v) if v else 0
            row_data.append("" if v is None else v)
        ws.append(row_data)

        dr = ws.max_row
        row_bg = _SURFACE0 if ri % 2 == 0 else _BG

        raw_color = course.get("color", "#89B4FA").lstrip("#").upper()
        # Guard against invalid hex
        if len(raw_color) != 6 or not all(c in "0123456789ABCDEFabcdef"
                                           for c in raw_color):
            raw_color = _ACCENT

        for ci, cell in enumerate(ws[dr]):
            cell.fill      = PatternFill("solid", fgColor=row_bg)
            cell.font      = Font(name="Arial", size=9, color=_FG)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 0))
            cell.border    = thin_border

        # Course Name: bold, course colour
        ws[dr][0].font = Font(name="Arial", size=9, bold=True, color=raw_color)
        # Credits: centered accent
        ws[dr][3].font = Font(name="Arial", size=9, bold=True, color=_ACCENT)
        ws[dr][3].alignment = Alignment(horizontal="center", vertical="center")
        # Exam Given: color-coded
        exam_val = ws[dr][4].value
        if exam_val == "TRUE":
            ws[dr][4].font = Font(name="Arial", size=9, bold=True, color=_GREEN)
        else:
            ws[dr][4].font = Font(name="Arial", size=9, color=_OVERLAY)
        ws[dr][4].alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[dr].height = 18

    # ── Totals row ────────────────────────────────────────────────────────────
    courses = semester.get("courses", [])
    total_credits = sum(c.get("credits", 0) for c in courses)
    done_credits  = sum(c.get("credits", 0) for c in courses if c.get("exam_given"))
    ws.append(["", f"TOTAL — {len(courses)} courses", "", total_credits,
               f"{done_credits} earned", "", "", "", "", "", ""])
    tr = ws.max_row
    for ci, cell in enumerate(ws[tr]):
        cell.fill   = PatternFill("solid", fgColor=_SURFACE0)
        cell.font   = Font(name="Arial", size=9, bold=True, color=_FG)
        cell.border = Border(top=Side(style="medium", color=_ACCENT))
        cell.alignment = Alignment(vertical="center")
    ws[tr][1].font = Font(name="Arial", size=9, bold=True, color=_ACCENT)
    ws[tr][3].font = Font(name="Arial", size=9, bold=True, color=_ACCENT)
    ws[tr][4].font = Font(name="Arial", size=9, bold=True, color=_GREEN)
    ws.row_dimensions[tr].height = 20

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_letter, width in zip("ABCDEFGHIJK",
                                 [32, 20, 22, 8, 11, 13, 10, 13, 10, 28, 12]):
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A4"   # freeze title + header

    wb.save(path)


def export_template_xlsx(path: str):
    """Write a pre-filled template the user can edit and re-import."""
    export_semester_xlsx(
        {
            "name":         "SuSe26",
            "display_name": "Summer Semester 2026",
            "courses":      _TEMPLATE_COURSES,
        },
        path,
    )


def import_semester_xlsx(path: str) -> dict:
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    # Semester meta from row 1
    sem_name = ""
    sem_disp = ""
    if rows and rows[0] and rows[0][0]:
        title_str = str(rows[0][0])
        m_short = re.search(r"short name:\s*(\S+)", title_str)
        if m_short:
            sem_name = m_short.group(1).strip()
        m_disp = re.search(r"^\s*(.+?)\s+·", title_str)
        if m_disp:
            sem_disp = m_disp.group(1).strip()

    # Find header row
    header_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == XLSX_HEADERS[0]:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(
            "Could not locate header row.\n"
            "Make sure the first column header is 'Course Name'."
        )

    col_map = {str(h).strip(): i
               for i, h in enumerate(rows[header_idx])
               if h is not None}

    courses = []
    for row in rows[header_idx + 1:]:
        if not row or not any(v for v in row if v is not None and str(v).strip()):
            continue
        str_row = [str(v) if v is not None else "" for v in row]
        course  = _parse_course_row(str_row, col_map)
        if course.get("name"):         # skip empty/totals rows
            courses.append(course)

    return {
        "name":         sem_name,
        "display_name": sem_disp,
        "courses":      courses,
        "exams":        [],
    }